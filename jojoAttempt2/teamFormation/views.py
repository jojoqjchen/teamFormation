from django.shortcuts import render
from django.shortcuts import redirect
from .models import csvUpload, numberOfDownloads # Import the models
from .forms import fileForm, colForm, teamSizeForm, projectFirstParamForm # Import the forms
import csv
from django.http import HttpResponse
import random
from django.contrib.auth.decorators import login_required # Easily define login requirements for views
from django.urls import reverse # Generate URLs using the name as defined in views.py
import openpyxl
import os
import xlwt
from teamFormationCode.script import team_formation_tool # Python script to generate teams
from teamFormationCode.project_first import project_first_teams # Python script to generate teams
from django.contrib.auth.models import User # Import the base User model
from django.db.models import Sum # To query the database and sum results
import pandas as pd

# Below: needed to output PDFs
import io
from django.http import FileResponse
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph


# Create your views here.

def home(request):
    downloads = numberOfDownloads.objects.all().aggregate(Sum('download'))
    users = User.objects.all().count()
    print(downloads['download__sum'],users)
    return render(request, 'index.html', {'downloads':downloads['download__sum'],'users':users})

# Step 1: Upload CSV File
@login_required
def uploadFile(request):

    instructions = 'Enter your CSV or Excel file containing the information about the individuals.'

    # If the form is filled
    if request.method == 'POST':
        # Validation may go here:
        # https://stackoverflow.com/questions/54403638/django-csv-file-validation-in-model-form-clean-method
        form = fileForm(request.POST, request.FILES) # Creating the form -> this allow to check for the correct extension

        query = request.POST.copy() # !IMPORTANT
        query.pop('csrfmiddlewaretoken') # Removing unwanted information
        algorithm = list(query.values()) # Get the answers provided for each of the columns in the initial form
        request.session['algorithm'] = algorithm[0]
        try:
            extension = os.path.splitext(str(request.FILES['csvFile']))[1]
            file = request.FILES['csvFile']
            data=[]

            ### TESTING EXTENSION
            if extension == '.xlsx':
                wb = openpyxl.load_workbook(file)
                # getting a particular sheet by name out of many sheets
                sheets = wb.sheetnames
                worksheet = wb[sheets[0]]

                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    data.append(row_data)

            elif extension == '.csv':
                # Getting the data from the form
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                for row in reader:
                    data.append(row)
            else: # Currently, return home template with a WARNING - Incorrect extension
                return render(request, 'team-formation/team-formation-tool.html', {'form': fileForm(), 'step': '1', 'warning': 'Incorrect extension. Please make sure to upload a CSV or an Excel file.', 'instructions': instructions})

            ### SAVING FORM DATA IN THE SESSION
            colNames = data.pop(0)
            request.session['colNames'] = colNames # Save column names in session
            request.session['data'] = data # Save the rest of the data in session

            # GOING TO THE NEXT STEP
            if algorithm[0] == '1':
                return redirect('/columns/') # Redirect to the next step -> will call pickColumns
            else:
                return redirect('/project-first-param/')
        except: # Currently, return home template with a WARNING - Incorrect extension

            return render(request, 'team-formation/team-formation-tool.html', {'form': fileForm(), 'step': '1', 'warning': 'Incorrect extension. Please make sure to upload a CSV or an Excel file.', 'instructions': instructions})

    # If the form is not filled -> we create it
    else:
        form = fileForm()

    return render(request, 'team-formation/team-formation-tool.html', {'form': form, 'step': '1', 'instructions': instructions})

# Step 2: Pick similar and different columns
@login_required
def pickColumns(request):

    instructions = 'Select the characteristics you want to optimize your teams on, or discard as many as you want.'
    colNames = list(request.session['colNames']) # list() may be unnecessary
    # print(colNames)
    answers = []

    for idx, col in enumerate(colNames):
        colNames[idx] = str(colNames[idx]+' ('+str(idx+1)+')') # Adding a # to each field -> in case there are two similar fields names

    # If the form is filled…
    if request.method == 'POST':

        # DOC: https://docs.djangoproject.com/en/3.2/ref/request-response/
        query = request.POST.copy() # !IMPORTANT
        query.pop('csrfmiddlewaretoken') # Removing unwanted information
        answers_raw = list(query.values()) # Get the answers provided for each of the columns in the initial form

        idxNumericCol = request.session['idxNumericCol']
        answers = colNames.copy()
        for i in range(len(colNames)):
            if i in idxNumericCol:
                idx = idxNumericCol.index(i)
                answers[i] = answers_raw[idx]
            else:
                answers[i] = "3"
        request.session['answers'] = answers
        print(answers)
        return redirect('/teamsize/')

    # Else, we need to create a dynamic form with the columns from the imported csv file
    else:
        colNameIsNumeric = [] # Name of columns that contain numbers
        idxNumericCol = []
        row = request.session['data'][0]
        # print("row", row)

        for i in range(0, len(colNames)): # For each column
            if row[i].isnumeric(): # Checking if the cells in the column contain numeric data
                colNameIsNumeric.append(colNames[i]) # If yes, then add the column name in
                idxNumericCol.append(i)
                # print("col", colNames[i])
            else:
                colNameIsNumeric.append('')
        request.session['idxNumericCol'] = idxNumericCol
        form = colForm(colNames,colNameIsNumeric) # See forms.py for further details

    return render(request, 'team-formation/team-formation-tool.html', {'form': form, 'step': '2', 'long': True, 'previous':"upload-teams", 'instructions': instructions})

@login_required
def projectFirstParam(request):

    instructions = 'Complete the characteristics of the projects.'

    # If the form is filled
    if request.method == 'POST':

        numberOfProjects = request.POST.get('numberOfProjects')
        numberOfChoices = request.POST.get('numberOfChoices')
        request.session['numberOfProjects'] = numberOfProjects
        request.session['numberOfChoices'] = numberOfChoices

        return redirect('/teamsize/')

    # If the form has not been filled yet
    else:

        form = projectFirstParamForm()

    return render(request, 'team-formation/team-formation-tool.html', {'form': form, 'step': '2', 'long': True, 'previous': "upload-teams", 'instructions': instructions})

 # Step 3: Enter team size
@login_required
def teamSize(request):

    instructions = 'Enter the ideal size for the teams.'

    # If the form is filled
    if request.method == 'POST':

        size = request.POST.get('size')
        request.session['size'] = size

        return render(request, 'team-formation/success.html')

    # If the form has not been filled yet
    else:

        form = teamSizeForm()
    previous = "columns" if request.session['algorithm'] == "1" else "projectFirstParam"

    return render(request, 'team-formation/team-formation-tool.html', {'form': form, 'step': '3', 'long': True, 'previous': previous, 'instructions': instructions})

# IDEA: Gather the two "download" views -> i.e. passing the format of the output in the argument
@login_required
def downloadResultCsv(request):

    response = HttpResponse(
        content_type='text/csv',
        headers = {'Content-Disposition': 'attachment; filename="team-formation-results.csv"'},
    )

    colNames = list(request.session['colNames'])
    data = request.session['data']
    size = request.session['size']
    answers = request.session['answers']
    algorithm = request.session['algorithm']
    numberOfProjects = request.session['numberOfProjects']
    numberOfChoices = request.session['numberOfChoices']

    if algorithm == '1':
        report = team_formation_tool(data,colNames,answers,int(size),False)
    else:
        report = project_first_teams(data,colNames, int(numberOfProjects), int(size), int(numberOfChoices), False)

    colNames.append('Team')
    writer = csv.writer(response)
    writer.writerow(colNames) # To update

    for i in range(0,report.shape[0]):
        writer.writerow(list(report.iloc[i,:]))

    user = numberOfDownloads.objects.get(user = request.user)
    user.download+=1
    user.save()

    return response

@login_required
def downloadResultXlsx(request):
    response = HttpResponse(
        content_type = 'application/ms-excel',
        headers = {'Content-Disposition': 'attachment; filename="team-formation-results.xls"'}
    )

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Teams')

    font_style = xlwt.XFStyle()
    font_style.font.bold = True

    colNames = list(request.session['colNames'])
    data = request.session['data']
    size = request.session['size']
    answers = list(request.session['answers'])
    algorithm = request.session['algorithm']
    numberOfProjects = request.session['numberOfProjects']
    numberOfChoices = request.session['numberOfChoices']

    if algorithm == '1':
        report = team_formation_tool(data,colNames,answers,int(size),False)
    else:
        report = project_first_teams(data,colNames, int(numberOfProjects), int(size), int(numberOfChoices), False)

    colNames.append('Team')
    nCol = len(colNames)

    for col_num in range(nCol):
        ws.write(0, col_num, colNames[col_num],font_style)

    font_style = xlwt.XFStyle()

    for i in range(0,report.shape[0]):
        for col_num in range(nCol):
            ws.write(i+1, col_num, str(report.iloc[i,col_num]), font_style)

    wb.save(response)

    user = numberOfDownloads.objects.get(user = request.user)
    user.download+=1
    user.save()

    return response

@login_required
def downloadResultPdf(request): #https://www.youtube.com/watch?v=_zkYICsIbXI&ab_channel=CryceTruly

    data = request.session['data']
    colNames = list(request.session['colNames'])
    answers = list(request.session['answers'])
    size = request.session['size']
    answers = list(request.session['answers'])
    algorithm = request.session['algorithm']

    if algorithm == '1':
        report = team_formation_tool(data,colNames,answers,int(size),False)
    else:
        numberOfProjects = request.session['numberOfProjects']
        numberOfChoices = request.session['numberOfChoices']
        report = project_first_teams(data,colNames, int(numberOfProjects), int(size), int(numberOfChoices), False)

    # NOTE: Because sometimes there are too many columns/values are too "long"
    # We need to limit both the max column size and the max number of columns
    cols = report.columns
    print(cols)
    if len(cols) > 7:
        data_pd = report.iloc[:,[0,1,2,3,4,5,-1]]
        data_pd = data_pd.astype('string')
        print(data_pd.info())
        data_pd = data_pd.apply(lambda x: x.str[:10])
        data_pd.columns = data_pd.columns.to_series().apply(lambda x: x[:10])
        dataWithTeams = [data_pd.columns]+data_pd.values.tolist()
    else:
        dataWithTeams = [cols] + report.values.tolist()

    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=letter)

    # container for the 'Flowable' objects
    elements = []

    t=Table(dataWithTeams)
    t.setStyle(TableStyle([('ALIGN',(1,1),(-2,-2),'RIGHT'),
    ('TEXTCOLOR',(0,0),(-1,0),colors.Color(253/255, 181/255, 21/255)),
    ('TEXTCOLOR', (0,1),(-1,-1), colors.Color(0, 50/255, 98/255)),
    ('VALIGN',(0,0),(0,-1),'MIDDLE'),
    ('ALIGN',(0,1),(-1,-1),'CENTER'),
    ('INNERGRID', (0,0), (-1,-1), 0.25, colors.black),
    ('BOX', (0,0), (-1,-1), 0.25, colors.black),
    ('FONT',(0,0),(-1,0),'Helvetica-Bold'),
    ]))

    # # First, second and third choices percentages
    # n = len(report)
    # print(n)
    # print(len(report[report['Team']==report['Choice 1']]))
    # perc_choice1 = len(report[report['Team']==report['Choice 1']])/n
    # perc_choice2 = len(report[report['Team']==report['Choice 2']])/n
    # perc_choice3 = len(report[report['Team']==report['Choice 3']])/n
    #
    # p1 = Paragraph("1st choice allocation: {}".format(perc_choice1), style=None)
    # p2 = Paragraph("2nd choice allocation: {}".format(perc_choice2), style=None)
    # p3 = Paragraph("3rd choice allocation: {}".format(perc_choice3), style=None)
    #

    elements.append(t)
    # elements.append(p1)
    # elements.append(p2)
    # elements.append(p3)

    doc.build(elements)

    buffer.seek(0)

    user = numberOfDownloads.objects.get(user = request.user)
    user.download+=1
    user.save()

    return FileResponse(buffer, as_attachment=True, filename='Team-Formation-Results.pdf')

### FURTHER TESTS

# test_session and test_delete Checks whether the browser accepted the cookie or not
def test_session(request):
    request.session.set_test_cookie()
    return HttpResponse("Testing session cookie")

def test_delete(request):
    if request.session.test_cookie_worked():
        request.session.delete_test_cookie()
        response = HttpResponse("Cookie test passed")
    else:
        response = HttpResponse("Cookie test failed")
    return response

# EXAMPLE views to manage session data
def save_session_data(request):
    request.session['data'] = request.GET.get('data')
    return HttpResponse("Session Data Saved")

def access_session_data(request):
    response = []
    if request.session.get('data'):
        response.append(request.session.get('data'))
    if not response:
        return HttpResponse("No Session Data")
    else:
        return HttpResponse(response)

def delete_session_data(request):
    try:
        del request.session['data']
    except KeyError:
        pass
    return HttpResponse("Session Data Cleared")
